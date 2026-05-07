from __future__ import annotations

import os
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from trading_bot.backtest.data import KiteHistoricalDataProvider, NseBhavcopyDataProvider, ZerodhaCredentials
from trading_bot.backtest.engine import run_backtest
from trading_bot.backtest.export import export_backtest_workbook
from trading_bot.backtest.models import BacktestConfig, StrategyConfig
from trading_bot.backtest.pdf_strategy import extract_pdf_text, infer_strategy_config
from trading_bot.backtest.universe import NIFTY_50_SYMBOLS
from trading_bot.config.settings import EXCEL_PATH
from trading_bot.excel.system_tab import read_capital_settings, append_trade_row, read_open_trades, write_trade_exit
from trading_bot.excel.workbook import WorkbookManager
from trading_bot.execution.paper_engine import (
    PaperSignal, ExitRecommendation,
    scan_signals, evaluate_exits, signal_to_trade_row,
)
from trading_bot.utils.logging import get_logger, _LOG_FILE

import hashlib
import json as _json

logger = get_logger(__name__)


BASE_DIR = Path(__file__).resolve().parents[2]
DEFAULT_PDF_PATH = Path(os.getenv("PDF_PATH", str(Path.home() / "Downloads" / "Swing_Trading_Framework_Investors_Way.pdf")))
DEFAULT_EXCEL_PATH = Path(os.getenv("EXCEL_PATH", str(EXCEL_PATH)))
CACHE_DIR = BASE_DIR / "trading_bot" / "data" / "cache"
RESULTS_DIR = BASE_DIR / "trading_bot" / "data" / "results"

STRATEGY_LABELS = {
    "sharique_swing":          "Sharique Shamsudheen — Swing Trading",
    "pullback_20dma":          "Pullback to 20 DMA",
    "rsi_2_mean_reversion":    "RSI-2 Mean Reversion  ★ 76-79% win rate",
    "momentum_30":             "Momentum 30",
    "pullback_50ema":          "Pullback to 50 EMA  ★ 55-60% win rate",
    "breakout_52w":            "52-Week High Breakout + Volume",
    "vishnu_trading_strategy": "Vishnu Trading Strategy  ★ 55-60% win rate",
}


def main() -> None:
    st.set_page_config(page_title="Swing Trading Backtester", layout="wide", page_icon="📈")

    _sidebar()

    title_col, restart_col = st.columns([8, 1])
    title_col.title("📈 Swing Trading Backtester")
    title_col.caption("Indian equities backtester — NSE Bhavcopy / Zerodha Kite · NIFTY 50 universe · Excel export")
    if restart_col.button("🔄 Restart", use_container_width=True, help="Clear all session state and restart the app"):
        _app_keys = [
            "trades_df", "equity_df", "strategy_dict", "initial_capital",
            "fetch_total", "fetch_fetched", "fetch_skipped",
            "_strategy", "_strategy_form_initialized", "_last_strategy_style",
            "data_source", "price_frames", "backtest_config",
            "comparison_results", "data_coverage_stats",
            "paper_price_frames", "paper_signals", "paper_exits",
        ]
        for key in _app_keys:
            st.session_state.pop(key, None)
        st.rerun()

    st.session_state["_strategy"] = _strategy_editor()
    st.divider()
    _backtest_tab()
    st.divider()
    _results_tab()
    st.divider()
    _paper_trading_tab()


# ─── Sidebar ────────────────────────────────────────────────────────────────

def _sidebar() -> None:
    st.sidebar.header("Configuration")

    st.sidebar.subheader("File Paths")
    pdf_path = Path(st.sidebar.text_input("Strategy PDF", value=str(DEFAULT_PDF_PATH)))
    excel_path = Path(st.sidebar.text_input("Planner Workbook", value=str(DEFAULT_EXCEL_PATH)))
    st.session_state["pdf_path"] = pdf_path
    st.session_state["excel_path"] = excel_path

    st.sidebar.subheader("Market Data")
    source = st.sidebar.selectbox(
        "Data source",
        options=["nse_bhavcopy", "zerodha"],
        index=0,
        format_func=lambda v: {"nse_bhavcopy": "NSE Bhavcopy (free)", "zerodha": "Zerodha Kite"}[v],
    )
    st.session_state["data_source"] = source

    if source == "zerodha":
        api_key = os.getenv("ZERODHA_API_KEY", "").strip()
        token = os.getenv("ZERODHA_ACCESS_TOKEN", "").strip()
        if api_key and token:
            st.sidebar.success("Zerodha credentials found")
        else:
            st.sidebar.error("ZERODHA_API_KEY / ZERODHA_ACCESS_TOKEN missing")

    st.sidebar.divider()
    _sidebar_data_library()
    st.sidebar.caption(f"Cache: `{CACHE_DIR}`")


# ─── Strategy tab ────────────────────────────────────────────────────────────

def _strategy_editor() -> StrategyConfig:
    st.header("⚙️ Strategy")
    pdf_path: Path = st.session_state.get("pdf_path", DEFAULT_PDF_PATH)
    raw_text = ""
    inferred = StrategyConfig()

    col_load, col_status = st.columns([3, 2])
    with col_load:
        st.subheader("Strategy Setup")
    with col_status:
        if pdf_path.exists():
            try:
                raw_text = extract_pdf_text(pdf_path)
                inferred = infer_strategy_config(raw_text)
                st.success(f"PDF loaded: {pdf_path.name}", icon="✅")
            except Exception as exc:
                st.warning(f"PDF extraction failed: {exc}", icon="⚠️")
        else:
            st.info("No PDF found — using preset defaults.", icon="ℹ️")

    with st.expander("📄 Extracted strategy text", expanded=False):
        st.text_area("Source text", value=raw_text, height=200, label_visibility="collapsed")

    _initialize_strategy_form(inferred)

    style_options = list(STRATEGY_LABELS.keys())
    current_style = st.session_state.get("strategy_style", inferred.strategy_style)
    if current_style not in style_options:
        current_style = style_options[0]

    strategy_style = st.selectbox(
        "Strategy style",
        options=style_options,
        index=style_options.index(current_style),
        key="strategy_style",
        format_func=lambda v: STRATEGY_LABELS.get(v, v),
    )
    _apply_strategy_style_preset(inferred, strategy_style)

    if strategy_style == "rsi_2_mean_reversion":
        st.info("**Highest win rate strategy (76–79%).** Larry Connors RSI-2 method — buy extreme short-term dips in stocks that are still in a long-term uptrend. Needs ~200 days of data.", icon="⭐")
    elif strategy_style == "rsi_pullback_uptrend":
        st.info("**65–70% win rate.** Buy when RSI cools to 38–55 (healthy pullback, not breakdown) while the stock is above its 50 EMA. Works best on NIFTY 50 large caps.", icon="📉")
    elif strategy_style == "consolidation_breakout":
        st.info("**55–58% win rate.** Narrow-range consolidation followed by a volume-backed breakout. Particularly effective on post-earnings consolidations on NSE.", icon="📦")
    elif strategy_style == "pullback_50ema":
        st.info("**55–60% win rate.** Simple and clean — stock in long-term uptrend (above 200 EMA) pulls back to 50 EMA and shows a reversal candle. Needs ~200 days of data.", icon="📊")
    elif strategy_style == "sharique_swing":
        st.info("**Approximation** of Sharique Shamsudheen's swing framework — 50 EMA trend filter, 20 EMA pullback, RSI 35–60 zone, bullish reversal, above-avg volume. 2:1 RR.", icon="👤")
    elif strategy_style == "vishnu_trading_strategy":
        st.info("**55–60% win rate · Profit Factor 2.2–2.5. Weekly timeframe.** Strong candlestick pattern (morning star / engulfing / piercing / three white soldiers) + MACD negative-zone crossover + volume above 20W avg + price above 12W & 26W EMA + support at 50W or 200W MA. Trailing SL activates at 2R — lets winners run while protecting capital.", icon="📈")

    # Build StrategyConfig from preset values stored in session state
    ss = st.session_state
    return StrategyConfig(
        name=ss.get("strategy_name", ""),
        strategy_style=strategy_style,
        timeframe=ss.get("strategy_timeframe", "day"),
        direction=ss.get("strategy_direction", "LONG"),
        trend_fast_ema=int(ss.get("trend_fast_ema", 50)),
        trend_slow_ema=int(ss.get("trend_slow_ema", 200)),
        signal_fast_ema=int(ss.get("signal_fast_ema", 20)),
        signal_slow_ema=int(ss.get("signal_slow_ema", 50)),
        volume_window=int(ss.get("volume_window", 20)),
        volume_multiplier=float(ss.get("volume_multiplier", 1.2)),
        support_window=int(ss.get("support_window", 20)),
        support_threshold_pct=float(ss.get("support_threshold_pct", 3.0)),
        stop_loss_pct=float(ss.get("stop_loss_pct", 8.0)),
        risk_reward_ratio=float(ss.get("risk_reward_ratio", 2.0)),
        max_holding_days=int(ss.get("max_holding_days", 20)),
        minimum_confidence=int(ss.get("minimum_confidence", 4)),
        require_macd_crossover=bool(ss.get("require_macd_crossover", True)),
        require_ema_alignment=bool(ss.get("require_ema_alignment", True)),
        require_support_bounce=bool(ss.get("require_support_bounce", True)),
        require_bullish_reversal=bool(ss.get("require_bullish_reversal", True)),
        require_volume_confirmation=bool(ss.get("require_volume_confirmation", True)),
        use_trailing_stop=bool(ss.get("use_trailing_stop", False)),
        trailing_trigger_r=float(ss.get("trailing_trigger_r", 2.0)),
        trailing_stop_pct=float(ss.get("trailing_stop_pct", 7.5)),
        notes=ss.get("strategy_notes", ""),
        raw_source_excerpt=raw_text[:4000],
    )


# ─── Backtest tab ────────────────────────────────────────────────────────────

def _backtest_tab() -> None:
    st.header("▶️ Backtest")
    excel_path: Path = st.session_state.get("excel_path", DEFAULT_EXCEL_PATH)
    workbook_settings = _load_workbook_defaults(excel_path)
    c1, c2, c3, c4, c5 = st.columns(5)
    start_date = c1.date_input("Start date", value=date.today() - timedelta(days=365))
    end_date = c2.date_input("End date", value=date.today())
    capital = c3.number_input("Capital (₹)", min_value=1000.0, value=float(workbook_settings["capital"]), step=1000.0)
    risk_pct = c4.number_input("Risk % per trade", min_value=0.1, max_value=10.0, value=float(workbook_settings["risk_pct"]), step=0.1)
    max_positions = c5.number_input("Max open trades", min_value=1, max_value=20, value=int(workbook_settings["num_trades"]))

    backtest_config = BacktestConfig(
        start_date=start_date,
        end_date=end_date,
        initial_capital=float(capital),
        risk_pct=float(risk_pct),
        max_open_positions=int(max_positions),
    )

    source = st.session_state.get("data_source", "nse_bhavcopy")
    provider = _build_provider(source)

    if source == "nse_bhavcopy":
        st.info("NSE Bhavcopy: downloads & caches one file per trading day. First multi-year run may take a few minutes.", icon="ℹ️")
    elif provider.available():
        st.success("Zerodha Kite credentials detected.", icon="✅")
    else:
        st.error("ZERODHA_API_KEY and ZERODHA_ACCESS_TOKEN are required for the Zerodha source.", icon="🚫")

    strategy: StrategyConfig = st.session_state.get("_strategy") or StrategyConfig()

    cache_key = _backtest_cache_key(strategy, backtest_config, source)
    cached = _load_backtest_cache(cache_key)

    # Warmup warning — shown before running
    warmup_days = max(strategy.trend_slow_ema, strategy.support_window)
    trading_days = int((backtest_config.end_date - backtest_config.start_date).days * 5 / 7)
    if trading_days < warmup_days + 5:
        months_needed = round((warmup_days + 20) / 21)
        st.warning(
            f"⚠️ **Not enough data for this strategy.** "
            f"The strategy needs **{warmup_days} days** of warmup (trend slow EMA = {strategy.trend_slow_ema}) "
            f"but your date range has ~{trading_days} trading days. "
            f"Extend the start date by at least **{months_needed} months** to get trades.",
            icon="⚠️",
        )

    if cached:
        st.info("Loaded from cache — parameters unchanged since last run.", icon="⚡")
        st.session_state.update(cached)

    # Show fetch status from previous run if available
    if "fetch_total" in st.session_state:
        fetched = st.session_state["fetch_fetched"]
        total = st.session_state["fetch_total"]
        skipped_list = st.session_state.get("fetch_skipped", [])
        color = "normal" if fetched == total else "off"
        st.metric("Companies fetched", f"{fetched} / {total}", delta=f"{total - fetched} skipped" if total - fetched else "all fetched", delta_color=color)
        if skipped_list:
            with st.expander(f"Skipped symbols ({len(skipped_list)})", expanded=False):
                st.write(", ".join(skipped_list))

    col_run, col_clear = st.columns([3, 1])
    run_clicked = col_run.button("▶️  Run Backtest", type="primary", use_container_width=True)
    clear_clicked = col_clear.button("🗑️  Clear cache", use_container_width=True)

    if clear_clicked:
        _clear_backtest_cache(cache_key)
        st.info("Cache cleared. Click Run Backtest to re-fetch.", icon="🗑️")

    if run_clicked:
        if not provider.available():
            st.error("Data source not configured correctly.")
        else:
            try:
                progress = st.progress(0, text="Fetching market data…")
                frames = {}
                skipped = []
                total = len(NIFTY_50_SYMBOLS)
                for i, symbol in enumerate(NIFTY_50_SYMBOLS):
                    try:
                        frames[symbol] = provider.get_daily_history(symbol, backtest_config.start_date, backtest_config.end_date)
                    except Exception as e:
                        skipped.append(symbol)
                        logger.warning("Skipping %s: %s", symbol, e)
                    progress.progress((i + 1) / total, text=f"Fetching {symbol}… ({i + 1}/{total})")
                st.session_state["fetch_total"] = total
                st.session_state["fetch_fetched"] = len(frames)
                st.session_state["fetch_skipped"] = skipped
                progress.progress(100, text="Running backtest…")
                if not frames:
                    st.error("No data could be fetched for any symbol. Try a different date range or data source.")
                    return
                trades_df, equity_df = run_backtest(frames, strategy, backtest_config)
                progress.empty()
                payload = {
                    "trades_df": trades_df,
                    "equity_df": equity_df,
                    "strategy_dict": strategy.to_dict(),
                    "initial_capital": backtest_config.initial_capital,
                }
                _save_backtest_cache(cache_key, payload)
                st.session_state.update(payload)
                # Store frames + config for comparison section; invalidate stale comparison
                st.session_state["price_frames"] = frames
                st.session_state["backtest_config"] = backtest_config
                st.session_state.pop("comparison_results", None)
                st.success(f"Backtest complete — {len(trades_df)} trades across {len(frames)} symbols.", icon="✅")
            except Exception as exc:
                st.exception(exc)


# ─── Results tab ─────────────────────────────────────────────────────────────

def _results_tab() -> None:
    st.header("📊 Results")
    trades_df: pd.DataFrame | None = st.session_state.get("trades_df")
    equity_df: pd.DataFrame | None = st.session_state.get("equity_df")
    strategy_dict: dict | None = st.session_state.get("strategy_dict")
    excel_path: Path = st.session_state.get("excel_path", DEFAULT_EXCEL_PATH)

    if trades_df is None or equity_df is None:
        st.info("Run a backtest first to see results here.", icon="ℹ️")
        return

    if trades_df.empty:
        strategy_dict = st.session_state.get("strategy_dict", {})
        warmup = max(strategy_dict.get("trend_slow_ema", 50), strategy_dict.get("support_window", 20))
        st.error(
            f"**0 trades generated.** Most likely cause: your date range is shorter than the strategy's "
            f"warmup period ({warmup} days for trend slow EMA = {strategy_dict.get('trend_slow_ema', '?')}). "
            f"Try extending the start date to at least {round((warmup + 20) / 21)} months back.",
            icon="🚫",
        )
        return

    initial_capital = float(st.session_state.get("initial_capital", 25000.0))
    metrics = _compute_metrics(trades_df, equity_df, initial_capital)

    c = st.columns(8)
    c[0].metric("Trades", metrics["trades"])
    c[1].metric("Win rate", f"{metrics['win_rate']:.1f}%")
    c[2].metric("Net P&L", f"₹{metrics['net_pnl']:,.0f}", delta=f"{metrics['net_pnl_pct']:.1f}%")
    c[3].metric("Avg P&L %", f"{metrics['avg_pnl_pct']:.2f}%")
    c[4].metric("Profit factor", f"{metrics['profit_factor']:.2f}")
    c[5].metric("Max drawdown", f"{metrics['max_drawdown_pct']:.1f}%")
    c[6].metric("Avg hold (days)", f"{metrics['avg_hold']:.0f}")
    c[7].metric("Final equity", f"₹{metrics['final_equity']:,.0f}")

    st.divider()

    # ── Charts ─────────────────────────────────────────────────────────────
    chart_col, pie_col = st.columns([3, 1])

    with chart_col:
        st.markdown("##### Equity Curve & Drawdown")
        _plot_equity_drawdown(equity_df, initial_capital)

    with pie_col:
        st.markdown("##### Win / Loss")
        _plot_win_loss_pie(trades_df)

    if not trades_df.empty:
        st.markdown("##### P&L by Symbol")
        _plot_symbol_pnl(trades_df)

    st.divider()

    # ── Trade table ────────────────────────────────────────────────────────
    st.markdown("##### Trade List")
    display_df = _style_trades(trades_df)
    st.dataframe(display_df, use_container_width=True, height=400)

    st.divider()

    # ── Export ─────────────────────────────────────────────────────────────
    col_exp, col_msg = st.columns([2, 3])
    with col_exp:
        if excel_path.exists():
            if st.button("💾  Export workbook copy", use_container_width=True):
                target = export_backtest_workbook(excel_path, trades_df, strategy_dict, RESULTS_DIR)
                st.success(f"Exported: `{target}`")
        else:
            st.warning(f"Workbook not found at `{excel_path}` — export unavailable.", icon="⚠️")

    st.divider()
    _comparison_section(current_strategy_dict=strategy_dict, initial_capital=initial_capital)


# ─── Strategy comparison ─────────────────────────────────────────────────────

_ALL_COMPARISON_STYLES = {
    "sharique_swing":          "Sharique Swing",
    "pullback_20dma":          "Pullback to 20 DMA",
    "rsi_2_mean_reversion":    "RSI-2 Mean Reversion",
    "momentum_30":             "Momentum 30",
    "pullback_50ema":          "Pullback to 50 EMA",
    "breakout_52w":            "52-Week Breakout",
    "vishnu_trading_strategy": "Vishnu Trading Strategy",
}


def _comparison_section(current_strategy_dict: dict, initial_capital: float) -> None:
    st.header("⚖️ Strategy Comparison")

    frames: dict | None = st.session_state.get("price_frames")
    backtest_config: BacktestConfig | None = st.session_state.get("backtest_config")

    if frames is None or backtest_config is None:
        st.info("Run a backtest first — comparison will appear here automatically.", icon="ℹ️")
        return

    # Run all 8 strategies if not already cached
    if "comparison_results" not in st.session_state:
        bar = st.progress(0, text="Running comparison across all strategies…")
        results = {}
        base = StrategyConfig()
        styles = list(_ALL_COMPARISON_STYLES.keys())
        for i, style in enumerate(styles):
            bar.progress((i + 1) / len(styles), text=f"Running {_ALL_COMPARISON_STYLES[style]}…")
            strategy = _preset_for_style(base, style)
            cache_key = _backtest_cache_key(strategy, backtest_config, st.session_state.get("data_source", "nse_bhavcopy"))
            cached = _load_backtest_cache(cache_key)
            if cached:
                trades_df = cached["trades_df"]
                equity_df = cached["equity_df"]
            else:
                try:
                    trades_df, equity_df = run_backtest(frames, strategy, backtest_config)
                    _save_backtest_cache(cache_key, {
                        "trades_df": trades_df,
                        "equity_df": equity_df,
                        "strategy_dict": strategy.to_dict(),
                        "initial_capital": initial_capital,
                    })
                except Exception as exc:
                    logger.warning("Comparison failed for %s: %s", style, exc)
                    continue
            results[style] = {
                "metrics": _compute_metrics(trades_df, equity_df, initial_capital),
                "equity_df": equity_df,
            }
        bar.empty()
        st.session_state["comparison_results"] = results

    results = st.session_state["comparison_results"]
    if not results:
        st.warning("No comparison results available.", icon="⚠️")
        return

    current_style = current_strategy_dict.get("strategy_style", "")

    # ── Summary table ──────────────────────────────────────────────────────
    st.markdown("##### Performance Summary")
    rows = []
    for style, data in results.items():
        m = data["metrics"]
        rows.append({
            "Strategy": ("★ " if style == current_style else "") + _ALL_COMPARISON_STYLES.get(style, style),
            "Trades": m["trades"],
            "Win Rate": f"{m['win_rate']:.1f}%",
            "Net P&L": f"₹{m['net_pnl']:,.0f}",
            "P&L %": f"{m['net_pnl_pct']:.1f}%",
            "Profit Factor": f"{m['profit_factor']:.2f}" if m["profit_factor"] != float("inf") else "∞",
            "Max DD": f"{m['max_drawdown_pct']:.1f}%",
            "Avg Hold": f"{m['avg_hold']:.0f}d",
            "_pnl": m["net_pnl"],
            "_style": style,
        })

    rows.sort(key=lambda r: float(r["Win Rate"].rstrip("%")), reverse=True)
    table_df = pd.DataFrame(rows).drop(columns=["_pnl", "_style"])
    st.dataframe(table_df, use_container_width=True, hide_index=True)

    # ── Multi-strategy equity curve ────────────────────────────────────────
    st.markdown("##### Equity Curves")
    fig = go.Figure()
    fig.add_hline(y=initial_capital, line_dash="dot", line_color="gray", annotation_text="Initial capital")
    for style, data in results.items():
        eq = data["equity_df"].copy()
        eq["date"] = pd.to_datetime(eq["date"])
        eq = eq.sort_values("date")
        is_current = style == current_style
        fig.add_trace(go.Scatter(
            x=eq["date"],
            y=eq["equity"],
            name=_ALL_COMPARISON_STYLES.get(style, style),
            line=dict(width=3 if is_current else 1.5),
            opacity=1.0 if is_current else 0.7,
        ))
    fig.update_layout(
        height=400,
        margin=dict(l=0, r=0, t=8, b=0),
        yaxis=dict(tickprefix="₹"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)


# ─── Chart helpers ────────────────────────────────────────────────────────────

def _plot_equity_drawdown(equity_df: pd.DataFrame, initial_capital: float) -> None:
    eq = equity_df.copy()
    eq["date"] = pd.to_datetime(eq["date"])
    eq = eq.sort_values("date")
    eq["peak"] = eq["equity"].cummax()
    eq["drawdown_pct"] = (eq["equity"] - eq["peak"]) / eq["peak"] * 100

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=eq["date"], y=eq["equity"],
        name="Equity", line=dict(color="#22c55e", width=2),
        fill="tozeroy", fillcolor="rgba(34,197,94,0.08)",
    ))
    fig.add_hline(y=initial_capital, line_dash="dot", line_color="gray", annotation_text="Initial capital")
    fig.add_trace(go.Bar(
        x=eq["date"], y=eq["drawdown_pct"],
        name="Drawdown %", marker_color="rgba(239,68,68,0.4)",
        yaxis="y2",
    ))
    fig.update_layout(
        height=340,
        margin=dict(l=0, r=0, t=8, b=0),
        yaxis=dict(title="Equity (₹)", tickprefix="₹"),
        yaxis2=dict(title="Drawdown %", overlaying="y", side="right", ticksuffix="%"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        hovermode="x unified",
    )
    st.plotly_chart(fig, use_container_width=True)


def _plot_win_loss_pie(trades_df: pd.DataFrame) -> None:
    if trades_df.empty:
        st.caption("No trades.")
        return
    wins = int((trades_df["pnl"] > 0).sum())
    losses = int((trades_df["pnl"] <= 0).sum())
    fig = go.Figure(go.Pie(
        labels=["Win", "Loss"],
        values=[wins, losses],
        marker_colors=["#22c55e", "#ef4444"],
        hole=0.45,
        textinfo="label+percent",
    ))
    fig.update_layout(height=260, margin=dict(l=0, r=0, t=8, b=0), showlegend=False)
    st.plotly_chart(fig, use_container_width=True)


def _plot_symbol_pnl(trades_df: pd.DataFrame) -> None:
    sym_pnl = (
        trades_df.groupby("symbol")["pnl"].sum()
        .sort_values()
        .reset_index()
    )
    colors = ["#22c55e" if v >= 0 else "#ef4444" for v in sym_pnl["pnl"]]
    fig = go.Figure(go.Bar(
        x=sym_pnl["pnl"], y=sym_pnl["symbol"],
        orientation="h",
        marker_color=colors,
        text=[f"₹{v:,.0f}" for v in sym_pnl["pnl"]],
        textposition="outside",
    ))
    fig.update_layout(
        height=max(260, len(sym_pnl) * 22),
        margin=dict(l=0, r=60, t=8, b=0),
        xaxis=dict(tickprefix="₹"),
    )
    st.plotly_chart(fig, use_container_width=True)


# ─── Metrics ─────────────────────────────────────────────────────────────────

def _compute_metrics(trades_df: pd.DataFrame, equity_df: pd.DataFrame, initial_capital: float) -> dict:
    if trades_df.empty:
        return dict(trades=0, win_rate=0, net_pnl=0, net_pnl_pct=0, avg_pnl_pct=0,
                    profit_factor=0, max_drawdown_pct=0, avg_hold=0, final_equity=initial_capital)

    trades = len(trades_df)
    wins = int((trades_df["pnl"] > 0).sum())
    net_pnl = float(trades_df["pnl"].sum())
    gross_profit = float(trades_df.loc[trades_df["pnl"] > 0, "pnl"].sum())
    gross_loss = abs(float(trades_df.loc[trades_df["pnl"] < 0, "pnl"].sum()))
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else float("inf")

    eq = equity_df.copy()
    eq["peak"] = eq["equity"].cummax()
    max_drawdown_pct = float(((eq["equity"] - eq["peak"]) / eq["peak"] * 100).min())

    return dict(
        trades=trades,
        win_rate=(wins / trades) * 100 if trades else 0,
        net_pnl=net_pnl,
        net_pnl_pct=(net_pnl / initial_capital) * 100,
        avg_pnl_pct=float(trades_df["pnl_pct"].mean()),
        profit_factor=profit_factor,
        max_drawdown_pct=max_drawdown_pct,
        avg_hold=float(trades_df["holding_days"].mean()),
        final_equity=initial_capital + net_pnl,
    )


def _style_trades(trades_df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "symbol", "entry_date", "exit_date", "holding_days",
        "entry_price", "exit_price", "stop_loss", "target_price",
        "quantity", "pnl", "pnl_pct", "confidence", "score",
        "market_trend", "entry_point", "avg_volume", "macd_crossover",
        "ema_crossover", "price_position", "bullish_reversal", "comments",
    ]
    display = trades_df[[c for c in cols if c in trades_df.columns]].copy()
    if "pnl" in display.columns:
        display["pnl"] = display["pnl"].map(lambda v: f"₹{v:,.2f}")
    if "pnl_pct" in display.columns:
        display["pnl_pct"] = display["pnl_pct"].map(lambda v: f"{v:.2f}%")
    return display


# ─── Strategy form helpers ────────────────────────────────────────────────────

def _initialize_strategy_form(inferred: StrategyConfig) -> None:
    if st.session_state.get("_strategy_form_initialized"):
        return
    preset = _preset_for_style(inferred, inferred.strategy_style)
    _set_strategy_form_state(preset)
    st.session_state["_last_strategy_style"] = preset.strategy_style
    st.session_state["_strategy_form_initialized"] = True


def _apply_strategy_style_preset(inferred: StrategyConfig, selected_style: str) -> None:
    if st.session_state.get("_last_strategy_style") == selected_style:
        return
    preset = _preset_for_style(inferred, selected_style)
    _set_strategy_form_state(preset)
    st.session_state["_last_strategy_style"] = selected_style
    st.rerun()


def _preset_for_style(base: StrategyConfig, style: str) -> StrategyConfig:
    preset = StrategyConfig(**base.to_dict())
    preset.strategy_style = style

    if style == "rsi_2_mean_reversion":
        preset.name = "RSI-2 Mean Reversion"
        preset.trend_fast_ema = 50
        preset.trend_slow_ema = 200
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.0
        preset.support_window = 20
        preset.support_threshold_pct = 3.0
        preset.stop_loss_pct = 3.0
        preset.risk_reward_ratio = 1.5
        preset.max_holding_days = 4
        preset.minimum_confidence = 3
        preset.require_macd_crossover = False
        preset.require_ema_alignment = True
        preset.require_support_bounce = False
        preset.require_bullish_reversal = True
        preset.require_volume_confirmation = False
        preset.notes = (
            "Larry Connors RSI-2 mean reversion. Backtested win rate: 76-79%. "
            "Entry: price above 200 EMA + RSI(2) < 10 + bullish reversal candle. "
            "Short hold (3-5 days). Tight stop 3%."
        )
    elif style == "rsi_pullback_uptrend":
        preset.name = "RSI Pullback in Uptrend"
        preset.trend_fast_ema = 50
        preset.trend_slow_ema = 200
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.0
        preset.support_window = 20
        preset.support_threshold_pct = 4.0
        preset.stop_loss_pct = 4.0
        preset.risk_reward_ratio = 2.0
        preset.max_holding_days = 8
        preset.minimum_confidence = 4
        preset.require_macd_crossover = False
        preset.require_ema_alignment = True
        preset.require_support_bounce = False
        preset.require_bullish_reversal = True
        preset.require_volume_confirmation = False
        preset.notes = (
            "RSI cools to 38-55 zone while stock is in uptrend. Backtested win rate: 65-70%. "
            "Works best on NIFTY 50 large caps (HDFC Bank, ICICI Bank, Reliance). "
            "Entry: above 50 EMA + near 20 EMA + RSI 38-55 + bullish reversal. RR 2:1."
        )
    elif style == "consolidation_breakout":
        preset.name = "Consolidation Breakout + Volume"
        preset.trend_fast_ema = 50
        preset.trend_slow_ema = 200
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.5
        preset.support_window = 20
        preset.support_threshold_pct = 3.0
        preset.stop_loss_pct = 4.0
        preset.risk_reward_ratio = 2.5
        preset.max_holding_days = 15
        preset.minimum_confidence = 3
        preset.require_macd_crossover = False
        preset.require_ema_alignment = True
        preset.require_support_bounce = False
        preset.require_bullish_reversal = False
        preset.require_volume_confirmation = True
        preset.notes = (
            "Stock consolidates in narrow range, then breaks out on 1.5x average volume. "
            "Backtested win rate: 55-58%. Particularly strong on post-earnings consolidations "
            "and sector breakouts in India. RR 2.5:1."
        )
    elif style == "pullback_50ema":
        preset.name = "Pullback to 50 EMA"
        preset.trend_fast_ema = 50
        preset.trend_slow_ema = 200
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.0
        preset.support_window = 20
        preset.support_threshold_pct = 3.0
        preset.stop_loss_pct = 5.0
        preset.risk_reward_ratio = 2.0
        preset.max_holding_days = 12
        preset.minimum_confidence = 4
        preset.require_macd_crossover = False
        preset.require_ema_alignment = True
        preset.require_support_bounce = False
        preset.require_bullish_reversal = True
        preset.require_volume_confirmation = False
        preset.notes = (
            "Price pulls back to 50 EMA while long-term trend (200 EMA) is intact. "
            "Backtested win rate: 55-60%. Simpler and more reliable than multi-condition setups. "
            "Entry: above 200 EMA + at/near 50 EMA + bullish reversal candle. RR 2:1."
        )
    elif style == "breakout_52w":
        preset.name = "52-Week High Breakout + Volume (PDF)"
        preset.trend_fast_ema = 50
        preset.trend_slow_ema = 200
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.5
        preset.support_window = 20
        preset.support_threshold_pct = 3.0
        preset.stop_loss_pct = 6.0
        preset.risk_reward_ratio = 2.5
        preset.max_holding_days = 20
        preset.minimum_confidence = 3
        preset.require_macd_crossover = False
        preset.require_ema_alignment = False
        preset.require_support_bounce = False
        preset.require_bullish_reversal = False
        preset.require_volume_confirmation = True
        preset.notes = "From PDF: Swing_Trading_Framework_Investors_Way. Entry at 52-week high breakout with above-average volume."
    elif style == "pullback_20dma":
        preset.name = "Mean-Reversion Pullback to 20 DMA (PDF)"
        preset.trend_fast_ema = 20
        preset.trend_slow_ema = 50
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.0
        preset.support_window = 20
        preset.support_threshold_pct = 2.0
        preset.stop_loss_pct = 5.0
        preset.risk_reward_ratio = 2.0
        preset.max_holding_days = 10
        preset.minimum_confidence = 4
        preset.require_macd_crossover = False
        preset.require_ema_alignment = True
        preset.require_support_bounce = True
        preset.require_bullish_reversal = True
        preset.require_volume_confirmation = False
        preset.notes = "From PDF: Swing_Trading_Framework_Investors_Way. Pullback to 20 DMA in uptrend with bullish reversal candle."
    elif style == "momentum_30":
        preset.name = "NIFTY 200 Momentum 30 (PDF)"
        preset.trend_fast_ema = 50
        preset.trend_slow_ema = 200
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.0
        preset.support_window = 20
        preset.support_threshold_pct = 3.0
        preset.stop_loss_pct = 8.0
        preset.risk_reward_ratio = 1.5
        preset.max_holding_days = 20
        preset.minimum_confidence = 4
        preset.require_macd_crossover = False
        preset.require_ema_alignment = False
        preset.require_support_bounce = False
        preset.require_bullish_reversal = False
        preset.require_volume_confirmation = False
        preset.notes = "From PDF: Swing_Trading_Framework_Investors_Way. Momentum 30 — top 30 NIFTY 200 stocks by 6/12-month momentum score."
    elif style == "sharique_swing":
        preset.name = "Sharique Shamsudheen — Swing Trading"
        preset.trend_fast_ema = 20
        preset.trend_slow_ema = 50
        preset.signal_fast_ema = 20
        preset.signal_slow_ema = 50
        preset.volume_window = 20
        preset.volume_multiplier = 1.2
        preset.support_window = 20
        preset.support_threshold_pct = 3.0
        preset.stop_loss_pct = 5.0
        preset.risk_reward_ratio = 2.0
        preset.max_holding_days = 10
        preset.minimum_confidence = 4
        preset.require_macd_crossover = False
        preset.require_ema_alignment = True
        preset.require_support_bounce = False
        preset.require_bullish_reversal = True
        preset.require_volume_confirmation = True
        preset.notes = (
            "Strategy inspired by Sharique Shamsudheen's swing trading framework. "
            "Entry: price above 50 EMA (trend), pullback to 20 EMA, RSI cooling to 35–60, "
            "bullish reversal candle, above-average volume. SL: 5% below entry. RR: 2:1."
        )
    elif style == "vishnu_trading_strategy":
        preset.name = "Vishnu Trading Strategy"
        preset.timeframe = "week"
        preset.trend_fast_ema = 50
        preset.trend_slow_ema = 200
        preset.signal_fast_ema = 12
        preset.signal_slow_ema = 26
        preset.macd_fast = 12
        preset.macd_slow = 26
        preset.macd_signal = 9
        preset.volume_window = 20
        preset.volume_multiplier = 1.2
        preset.support_window = 20
        preset.support_threshold_pct = 5.0
        preset.stop_loss_pct = 5.0
        preset.risk_reward_ratio = 2.0
        preset.max_holding_days = 16
        preset.minimum_confidence = 5
        preset.require_macd_crossover = True
        preset.require_ema_alignment = True
        preset.require_support_bounce = False
        preset.require_bullish_reversal = True
        preset.require_volume_confirmation = True
        preset.use_trailing_stop = True
        preset.trailing_trigger_r = 2.0
        preset.trailing_stop_pct = 7.5
        preset.notes = (
            "Weekly-timeframe multi-factor strategy. "
            "Entry: strong bullish candlestick pattern + MACD negative-zone crossover "
            "+ volume above 20W avg + price above 12W & 26W EMA + support at 50W/200W MA. "
            "Trailing SL: moves to breakeven at 2R, then trails 7.5% below the highest price."
        )

    return preset


def _set_strategy_form_state(config: StrategyConfig) -> None:
    st.session_state["strategy_name"] = config.name
    st.session_state["strategy_direction"] = config.direction
    st.session_state["strategy_timeframe"] = config.timeframe
    st.session_state["trend_fast_ema"] = config.trend_fast_ema
    st.session_state["trend_slow_ema"] = config.trend_slow_ema
    st.session_state["signal_fast_ema"] = config.signal_fast_ema
    st.session_state["signal_slow_ema"] = config.signal_slow_ema
    st.session_state["volume_window"] = config.volume_window
    st.session_state["volume_multiplier"] = float(config.volume_multiplier)
    st.session_state["support_window"] = config.support_window
    st.session_state["support_threshold_pct"] = float(config.support_threshold_pct)
    st.session_state["stop_loss_pct"] = float(config.stop_loss_pct)
    st.session_state["risk_reward_ratio"] = float(config.risk_reward_ratio)
    st.session_state["max_holding_days"] = config.max_holding_days
    st.session_state["minimum_confidence"] = config.minimum_confidence
    st.session_state["require_macd_crossover"] = config.require_macd_crossover
    st.session_state["require_ema_alignment"] = config.require_ema_alignment
    st.session_state["require_support_bounce"] = config.require_support_bounce
    st.session_state["require_bullish_reversal"] = config.require_bullish_reversal
    st.session_state["require_volume_confirmation"] = config.require_volume_confirmation
    st.session_state["use_trailing_stop"] = config.use_trailing_stop
    st.session_state["trailing_trigger_r"] = float(config.trailing_trigger_r)
    st.session_state["trailing_stop_pct"] = float(config.trailing_stop_pct)
    st.session_state["strategy_notes"] = config.notes or ""


# ─── Backtest cache ───────────────────────────────────────────────────────────

def _backtest_cache_key(strategy: StrategyConfig, config: BacktestConfig, source: str) -> str:
    payload = {
        "strategy": strategy.to_dict(),
        "start": str(config.start_date),
        "end": str(config.end_date),
        "capital": config.initial_capital,
        "risk_pct": config.risk_pct,
        "max_positions": config.max_open_positions,
        "source": source,
    }
    digest = hashlib.md5(_json.dumps(payload, sort_keys=True, default=str).encode()).hexdigest()
    return digest


def _cache_dir() -> Path:
    path = RESULTS_DIR / "backtest_cache"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _load_backtest_cache(key: str) -> dict | None:
    base = _cache_dir() / key
    trades_path = base.with_suffix(".trades.csv")
    equity_path = base.with_suffix(".equity.csv")
    meta_path = base.with_suffix(".meta.json")
    if not (trades_path.exists() and equity_path.exists() and meta_path.exists()):
        return None
    try:
        trades_df = pd.read_csv(trades_path, parse_dates=["entry_date", "exit_date", "signal_date"])
        equity_df = pd.read_csv(equity_path, parse_dates=["date"])
        meta = _json.loads(meta_path.read_text())
        return {"trades_df": trades_df, "equity_df": equity_df, **meta}
    except Exception:
        return None


def _save_backtest_cache(key: str, payload: dict) -> None:
    base = _cache_dir() / key
    payload["trades_df"].to_csv(base.with_suffix(".trades.csv"), index=False)
    payload["equity_df"].to_csv(base.with_suffix(".equity.csv"), index=False)
    meta = {k: v for k, v in payload.items() if k not in {"trades_df", "equity_df"}}
    base.with_suffix(".meta.json").write_text(_json.dumps(meta, default=str))


def _clear_backtest_cache(key: str) -> None:
    base = _cache_dir() / key
    for suffix in (".trades.csv", ".equity.csv", ".meta.json"):
        p = base.with_suffix(suffix)
        if p.exists():
            p.unlink()


# ─── Log viewer ──────────────────────────────────────────────────────────────

def _render_log_tail(lines: int = 60) -> None:
    if not _LOG_FILE.exists():
        st.caption("No log file yet.")
        return
    with open(_LOG_FILE, "r", encoding="utf-8") as f:
        tail = f.readlines()[-lines:]
    st.code("".join(tail), language=None)


# ─── Data library ────────────────────────────────────────────────────────────

def _sidebar_data_library() -> None:
    st.sidebar.subheader("📦 Data Library")
    provider = NseBhavcopyDataProvider(cache_dir=CACHE_DIR)

    # Cache stats in session state — recompute only when explicitly refreshed
    if "data_coverage_stats" not in st.session_state:
        st.session_state["data_coverage_stats"] = provider.get_coverage_stats(years=5)
    stats = st.session_state["data_coverage_stats"]

    cached = stats["cached"]
    holidays = stats.get("holidays", 0)
    total = stats["expected"]
    missing_count = len(stats["missing"])
    accounted = cached + holidays
    pct = accounted / total * 100 if total else 0.0

    icon = "🟢" if pct >= 99 else ("🟡" if pct >= 60 else "🔴")

    st.sidebar.markdown(f"{icon} **{accounted:,} / {total:,} days**")
    st.sidebar.caption(f"{cached:,} trading days · {holidays:,} confirmed holidays")
    st.sidebar.progress(min(int(pct), 100), text=f"{pct:.0f}% covered")

    if stats["earliest"] and stats["latest"]:
        st.sidebar.caption(f"{stats['earliest']} → {stats['latest']}")

    if missing_count == 0:
        st.sidebar.success("All data accounted for", icon="✅")
    else:
        st.sidebar.caption(f"{missing_count} dates not yet fetched")
        with st.sidebar.expander(f"Missing dates ({missing_count})", expanded=False):
            st.write(", ".join(str(d) for d in stats["missing"]))
        if st.sidebar.button("⬇️ Download missing data", use_container_width=True):
            _run_bulk_download(provider, stats["missing"])

    # Symbol index
    st.sidebar.divider()
    st.sidebar.subheader("🗂️ Symbol Index")
    idx = provider.get_symbol_index_stats(NIFTY_50_SYMBOLS)
    if idx["built"] == idx["total"] and idx["unpivoted_dates"] == 0:
        st.sidebar.success(f"All {idx['total']} symbols indexed", icon="⚡")
    else:
        if idx["built"] < idx["total"]:
            st.sidebar.caption(f"{idx['built']}/{idx['total']} symbols built")
        if idx["unpivoted_dates"] > 0:
            st.sidebar.caption(f"{idx['unpivoted_dates']} new dates not yet indexed")
        if st.sidebar.button("🔨 Build Symbol Index", use_container_width=True,
                             help="One-time pass: reads each Bhavcopy file once, builds per-symbol CSVs for instant backtest loads"):
            _run_pivot(provider)

    # Log viewer
    with st.sidebar.expander("🪵 Recent logs", expanded=False):
        _render_log_tail()


def _run_bulk_download(provider: NseBhavcopyDataProvider, missing: list) -> None:
    if not missing:
        st.sidebar.info("Nothing to download.")
        return
    bar = st.sidebar.progress(0, text="Starting download…")
    status = st.sidebar.empty()

    def on_progress(done: int, total: int, current_date) -> None:
        bar.progress(int(done / total * 100), text=f"{current_date}  ({done}/{total})")
        status.caption(f"Last fetched: {current_date}")

    error_msg = None
    try:
        downloaded, skipped = provider.download_bulk(missing, on_progress=on_progress)
    except Exception as exc:
        error_msg = str(exc)
        downloaded, skipped = 0, 0

    bar.empty()
    status.empty()

    if error_msg:
        st.sidebar.error(f"Download interrupted: {error_msg}")
    else:
        st.session_state.pop("data_coverage_stats", None)
        st.sidebar.success(f"Done — {downloaded} files downloaded, {skipped} holidays/timeouts skipped.")
        if downloaded > 0:
            _run_pivot(provider)
        else:
            st.rerun()


def _run_pivot(provider: NseBhavcopyDataProvider) -> None:
    idx = provider.get_symbol_index_stats(NIFTY_50_SYMBOLS)
    total_dates = idx["unpivoted_dates"]
    if total_dates == 0:
        st.sidebar.info("Symbol index already up to date.")
        st.rerun()
        return

    bar = st.sidebar.progress(0, text="Building symbol index…")
    status = st.sidebar.empty()

    def on_progress(done: int, total: int, current_date) -> None:
        bar.progress(int(done / total * 100), text=f"Indexing {current_date}  ({done}/{total})")
        status.caption(f"Processing: {current_date}")

    try:
        processed = provider.pivot_bhavcopy_to_symbols(NIFTY_50_SYMBOLS, on_progress=on_progress)
    except Exception as exc:
        bar.empty()
        status.empty()
        st.sidebar.error(f"Pivot failed: {exc}")
        return

    bar.empty()
    status.empty()
    st.sidebar.success(f"Symbol index built — {processed} dates processed.")
    st.rerun()


# ─── Shared helpers ───────────────────────────────────────────────────────────

def _load_workbook_defaults(excel_path: Path) -> dict:
    import openpyxl
    defaults = {"capital": 25000.0, "risk_pct": 2.0, "num_trades": 5}
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        defaults.update(read_capital_settings(wb))
    return defaults


def _build_provider(source: str):
    if source == "zerodha":
        return KiteHistoricalDataProvider(
            cache_dir=CACHE_DIR,
            credentials=ZerodhaCredentials(
                api_key=os.getenv("ZERODHA_API_KEY", "").strip(),
                access_token=os.getenv("ZERODHA_ACCESS_TOKEN", "").strip(),
            ),
        )
    return NseBhavcopyDataProvider(cache_dir=CACHE_DIR)


# ─── Paper Trading tab ────────────────────────────────────────────────────────

def _paper_trading_tab() -> None:
    st.header("📋 Paper Trading")
    st.caption("Evaluate today's signals against the selected strategy, enter trades and record exits in the Swing Planner.")

    excel_path: Path = st.session_state.get("excel_path", DEFAULT_EXCEL_PATH)
    strategy: StrategyConfig = st.session_state.get("_strategy") or StrategyConfig()
    source = st.session_state.get("data_source", "nse_bhavcopy")
    provider = _build_provider(source)
    workbook_settings = _load_workbook_defaults(excel_path)

    capital = float(workbook_settings["capital"])
    risk_pct = float(workbook_settings["risk_pct"])
    max_positions = int(workbook_settings["num_trades"])

    # ── Open trades ──────────────────────────────────────────────────────────
    st.subheader("📂 Open Trades")

    import openpyxl
    open_trades: list[dict] = []
    if excel_path.exists():
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            open_trades = read_open_trades(wb)
        except Exception as exc:
            st.warning(f"Could not read Swing Planner: {exc}")
    else:
        st.info(f"Swing Planner not found at `{excel_path}`. Create it first or update the path in the sidebar.", icon="ℹ️")

    # Split into NIFTY 50 universe vs. outside
    universe_set = set(NIFTY_50_SYMBOLS)
    open_in_universe = [t for t in open_trades if t["stock"] in universe_set]
    open_outside_universe = [t for t in open_trades if t["stock"] not in universe_set]

    open_slots = max(0, max_positions - len(open_trades))

    # Auto-fetch latest prices for open trade symbols so exits show immediately on page load
    if open_in_universe and "paper_price_frames" not in st.session_state:
        scan_start = date.today() - timedelta(days=300)
        frames: dict = {}
        with st.spinner("Loading latest prices for open trades…"):
            for t in open_in_universe:
                sym = t["stock"]
                try:
                    frames[sym] = provider.get_daily_history(sym, scan_start, date.today())
                except Exception:
                    pass
        if frames:
            st.session_state["paper_price_frames"] = frames

    if not open_trades:
        st.info("No open trades in the Swing Planner.", icon="📭")
    else:
        # Warn about any outside-universe symbols
        if open_outside_universe:
            names = ", ".join(t["stock"] for t in open_outside_universe)
            st.warning(
                f"**{names}** {'is' if len(open_outside_universe) == 1 else 'are'} not in the NIFTY 50 universe — "
                "exit recommendations cannot be generated automatically. "
                "Update the exit price manually in the Swing Planner.",
                icon="⚠️",
            )

        paper_frames: dict | None = st.session_state.get("paper_price_frames")
        exits: list[ExitRecommendation] = []
        if paper_frames:
            exits = evaluate_exits(open_in_universe, paper_frames, strategy, date.today())

        exit_map = {e.symbol: e for e in exits}
        rows = []
        for t in open_trades:
            sym = t["stock"]
            ex = exit_map.get(sym)
            in_universe = sym in universe_set
            rows.append({
                "Symbol": sym,
                "Entry Date": t["entry_date"],
                "Entry ₹": t["entry"],
                "SL ₹": t["stop_loss"],
                "Target ₹": t["target"],
                "Qty": t["quantity"],
                "Latest ₹": f"₹{ex.latest_close:,.2f}" if ex else "—",
                "P&L": f"{'▲' if ex and ex.pnl >= 0 else '▼'} ₹{ex.pnl:,.0f} ({ex.pnl_pct:+.1f}%)" if ex else "—",
                "Days": ex.holding_days if ex else "—",
                "Action": (
                    {"HOLD": "🟡 HOLD", "TARGET": "🟢 EXIT — Target hit",
                     "STOP_LOSS": "🔴 EXIT — Stop loss", "MAX_DAYS": "🟠 EXIT — Max days"}.get(ex.reason, "—")
                    if ex else ("⚠️ Outside NIFTY 50" if not in_universe else "—")
                ),
            })

        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        exits_to_action = [e for e in exits if e.reason != "HOLD"]
        if exits_to_action and excel_path.exists():
            st.warning(f"**{len(exits_to_action)} trade(s) should be closed** based on latest prices.", icon="⚠️")
            if st.button(f"✅ Execute {len(exits_to_action)} Exit(s) in Swing Planner", type="primary"):
                try:
                    with WorkbookManager(excel_path) as wbm:
                        for ex in exits_to_action:
                            write_trade_exit(wbm.workbook, ex.row, ex.latest_close, ex.latest_date.isoformat())
                    st.success(f"Exits recorded for: {', '.join(e.symbol for e in exits_to_action)}", icon="✅")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Failed to write exits: {exc}")

    st.metric("Open slots", f"{open_slots} / {max_positions}", delta=f"{len(open_trades)} currently open", delta_color="off")

    # ── Signal scan ──────────────────────────────────────────────────────────
    st.subheader("🔍 Signal Scanner")
    st.caption(f"Strategy: **{strategy.name}** · Style: `{strategy.strategy_style}` · Min confidence: {strategy.minimum_confidence}/7")

    scan_days = st.slider("Lookback days for scan", min_value=100, max_value=500, value=300, step=50,
                          help="How many days of history to load for indicator calculation")

    col_scan, col_clear = st.columns([3, 1])
    scan_clicked = col_scan.button("🔍 Scan for Signals", type="primary", use_container_width=True)
    if col_clear.button("🗑️ Clear scan", use_container_width=True):
        st.session_state.pop("paper_price_frames", None)
        st.session_state.pop("paper_signals", None)
        st.session_state.pop("paper_exits", None)
        st.rerun()

    if scan_clicked:
        if not provider.available():
            st.error("Data source not configured. Check sidebar settings.")
        else:
            scan_start = date.today() - timedelta(days=scan_days)
            scan_end = date.today()
            progress = st.progress(0, text="Loading market data for scan…")
            frames = {}
            for i, symbol in enumerate(NIFTY_50_SYMBOLS):
                try:
                    frames[symbol] = provider.get_daily_history(symbol, scan_start, scan_end)
                except Exception:
                    pass
                progress.progress((i + 1) / len(NIFTY_50_SYMBOLS), text=f"Loading {symbol}… ({i+1}/{len(NIFTY_50_SYMBOLS)})")

            progress.progress(100, text="Evaluating signals…")
            signals = scan_signals(frames, strategy, capital, risk_pct, open_slots=-1)
            st.session_state["paper_price_frames"] = frames
            st.session_state["paper_signals"] = signals
            # Refresh exit recommendations with new data
            if open_trades:
                st.session_state["paper_exits"] = evaluate_exits(open_trades, frames, strategy, date.today())
            progress.empty()
            st.rerun()

    signals: list[PaperSignal] = st.session_state.get("paper_signals", [])

    if not signals and "paper_price_frames" in st.session_state:
        st.info("No signals match the strategy criteria today.", icon="🔎")

    if signals:
        st.success(f"**{len(signals)} signal(s)** found — {open_slots} open slot(s) available.", icon="📡")

        sig_rows = []
        for s in signals:
            sig_rows.append({
                "Symbol": s.symbol,
                "Date": s.signal_date,
                "Close ₹": f"₹{s.close:,.2f}",
                "SL ₹": f"₹{s.stop_loss:,.2f}",
                "Target ₹": f"₹{s.target:,.2f}",
                "R:R": f"1:{strategy.risk_reward_ratio:.1f}",
                "Qty": s.quantity,
                "Score": f"{s.score}/7",
                "Confidence": s.confidence,
                "Trend": s.market_trend,
                "Volume": s.avg_volume,
                "MACD": s.macd_crossover,
                "Reversal": s.bullish_reversal,
            })
        st.dataframe(pd.DataFrame(sig_rows), use_container_width=True, hide_index=True)

        # Only allow entering up to available open slots
        tradeable = signals[:open_slots] if open_slots > 0 else []
        if open_slots <= 0:
            st.warning("No open slots — close existing trades before entering new ones.", icon="🚫")
        elif excel_path.exists():
            enter_count = st.number_input(
                "How many top signals to enter?",
                min_value=1, max_value=len(tradeable),
                value=min(len(tradeable), open_slots),
            )
            if st.button(f"▶️ Enter {enter_count} Trade(s) in Swing Planner", type="primary"):
                try:
                    entered = []
                    with WorkbookManager(excel_path) as wbm:
                        for sig in tradeable[:enter_count]:
                            row_num = append_trade_row(wbm.workbook, signal_to_trade_row(sig))
                            entered.append(f"{sig.symbol} @ ₹{sig.close:,.2f} (row {row_num})")
                    st.success(f"Entered {len(entered)} trade(s):\n" + "\n".join(f"• {e}" for e in entered), icon="✅")
                    # Clear scan so open trades refresh on next rerun
                    st.session_state.pop("paper_signals", None)
                    st.session_state.pop("paper_price_frames", None)
                    st.rerun()
                except Exception as exc:
                    st.error(f"Failed to write trades: {exc}")
        else:
            st.warning(f"Swing Planner not found at `{excel_path}` — trades cannot be entered.", icon="⚠️")
