from __future__ import annotations

import os
from datetime import date, timedelta
from pathlib import Path

import dash
import dash_bootstrap_components as dbc
import openpyxl
import pandas as pd
import plotly.graph_objects as go
from dash import Input, Output, callback, dcc, html
from dash import dash_table

from trading_bot.backtest.data import NseBhavcopyDataProvider
from trading_bot.backtest.models import StrategyConfig
from trading_bot.backtest.universe import NIFTY_50_SYMBOLS
from trading_bot.config.settings import EXCEL_PATH
from trading_bot.excel.system_tab import read_capital_settings, read_open_trades, write_trade_exit
from trading_bot.excel.workbook import WorkbookManager
from trading_bot.execution.paper_engine import evaluate_exits

dash.register_page(__name__, path="/", name="Home")

CACHE_DIR = Path(__file__).resolve().parents[4] / "trading_bot" / "data" / "cache"
_EXCEL = Path(os.getenv("EXCEL_PATH", str(EXCEL_PATH)))


def _balances() -> tuple[float, float, int]:
    """Return (capital, cash_available, open_count)."""
    try:
        wb = openpyxl.load_workbook(_EXCEL, data_only=True)
        s = read_capital_settings(wb)
        trades = read_open_trades(wb)
        capital = s["capital"]
        deployed = sum(
            float(t.get("entry") or 0) * int(t.get("quantity") or 0)
            for t in trades
        )
        return capital, max(0.0, capital - deployed), len(trades)
    except Exception:
        return 100000.0, 100000.0, 0


def _portfolio_chart(period: str) -> go.Figure:
    """Build equity curve from closed paper trades."""
    fig = go.Figure()
    try:
        wb = openpyxl.load_workbook(_EXCEL, data_only=True)
        ws = wb["System"]
        rows = []
        for row in range(19, 50):
            stock = ws.cell(row=row, column=2).value
            exit_price = ws.cell(row=row, column=18).value
            entry = ws.cell(row=row, column=14).value
            qty = ws.cell(row=row, column=16).value
            exit_date_val = ws.cell(row=row, column=19).value
            if stock and exit_price and entry and qty:
                rows.append({
                    "date": pd.to_datetime(exit_date_val) if exit_date_val else pd.Timestamp.now(),
                    "pnl": (float(exit_price) - float(entry)) * int(qty),
                })
        if rows:
            df = pd.DataFrame(rows).sort_values("date")
            df["equity"] = 100000 + df["pnl"].cumsum()
            cutoff = {
                "1D": date.today() - timedelta(days=1),
                "1M": date.today() - timedelta(days=30),
                "1Y": date.today() - timedelta(days=365),
            }.get(period)
            if cutoff:
                df = df[df["date"].dt.date >= cutoff]
            fig.add_trace(go.Scatter(
                x=df["date"], y=df["equity"],
                mode="lines", fill="tozeroy",
                line=dict(color="#F0B429", width=2),
                fillcolor="rgba(240,180,41,0.15)",
                hovertemplate="₹%{y:,.0f}<extra></extra>",
            ))
    except Exception:
        pass

    capital = _balances()[0]
    fig.add_hline(y=capital, line_dash="dot", line_color="#9CA3AF", line_width=1)
    fig.update_layout(
        height=200, margin=dict(l=0, r=0, t=8, b=8),
        paper_bgcolor="white", plot_bgcolor="white",
        xaxis=dict(showgrid=False, showline=False, zeroline=False),
        yaxis=dict(showgrid=True, gridcolor="#F3F4F6", zeroline=False, tickprefix="₹", tickformat=",.0f"),
        showlegend=False, hovermode="x unified",
    )
    return fig


def layout() -> html.Div:
    capital, cash, open_count = _balances()
    daily_pnl = 0.0

    return html.Div([
        # Info banner
        html.Div(
            [html.I(className="bi bi-info-circle me-2"),
             "You are on Paper Trading, no real money is being used."],
            className="info-banner",
        ),

        # Portfolio card
        html.Div(className="dash-card", children=[
            html.Div(className="dash-card-header", children=[
                html.Div([html.I(className="bi bi-graph-up me-2"), "Your Portfolio"], className="dash-card-title"),
                html.Div([
                    dbc.Button("1D", id="tf-1d", size="sm", className="time-filter-btn me-1"),
                    dbc.Button("1M", id="tf-1m", size="sm", className="time-filter-btn me-1"),
                    dbc.Button("1Y", id="tf-1y", size="sm", className="time-filter-btn me-1"),
                    dbc.Button("All", id="tf-all", size="sm", className="time-filter-btn active"),
                    dbc.Button(html.I(className="bi bi-arrow-clockwise"), id="tf-refresh", size="sm",
                               className="time-filter-btn ms-1"),
                ], style={"display": "flex", "alignItems": "center"}),
            ]),
            html.Div(f"$ {capital:,.2f}", className="portfolio-value"),
            html.Div(f"{date.today().strftime('%b %d, %Y')}", className="portfolio-date"),
            dcc.Graph(id="portfolio-chart", figure=_portfolio_chart("All"),
                      config={"displayModeBar": False}),
            dcc.Store(id="tf-store", data="All"),
        ]),

        # Balances row
        dbc.Row([
            dbc.Col(html.Div([
                html.Div("Buying Power", className="balance-label"),
                html.Div(f"₹{capital:,.2f}", className="balance-value"),
            ], className="balance-card"), width=4),
            dbc.Col(html.Div([
                html.Div("Cash", className="balance-label"),
                html.Div(f"₹{cash:,.2f}", className="balance-value"),
            ], className="balance-card"), width=4),
            dbc.Col(html.Div([
                html.Div("Daily Change", className="balance-label"),
                html.Div(
                    f"{'▲' if daily_pnl >= 0 else '▼'} ₹{abs(daily_pnl):,.2f}",
                    className="balance-value",
                    style={"color": "#00A852" if daily_pnl >= 0 else "#E03131"},
                ),
            ], className="balance-card"), width=4),
        ], className="mb-3 g-3"),

        # Open Positions
        html.Div(className="dash-card", children=[
            html.Div(className="dash-card-header", children=[
                html.Div([html.I(className="bi bi-briefcase me-2"), "Open Positions"], className="dash-card-title"),
                html.Span(f"{open_count} open",
                          style={"fontSize": "12px", "color": "#6B7280",
                                 "background": "#F3F4F6", "padding": "2px 8px", "borderRadius": "10px"}),
            ]),
            html.Div(id="positions-table-div", children=_positions_table()),
            html.Div(id="exit-msg"),
            dbc.Button("✅ Execute Exits", id="exec-exits-btn", className="btn-alpaca mt-2",
                       style={"display": "none"}, n_clicks=0),
        ]),

        # Recent Trades
        html.Div(className="dash-card", children=[
            html.Div(className="dash-card-header", children=[
                html.Div([html.I(className="bi bi-clock-history me-2"), "Recent Trades"], className="dash-card-title"),
            ]),
            html.Div(id="recent-trades-div", children=_recent_trades_table()),
        ]),

        dcc.Interval(id="home-refresh", interval=60_000, n_intervals=0),
    ], className="page-content")


def _positions_table() -> html.Div:
    try:
        wb = openpyxl.load_workbook(_EXCEL, data_only=True)
        trades = read_open_trades(wb)
    except Exception:
        trades = []

    if not trades:
        return html.Div("No open positions. Place some trades to see this table populate.",
                        style={"color": "#9CA3AF", "fontSize": "13px", "padding": "12px 0"})

    rows = [{
        "Asset": t["stock"],
        "Entry Date": t["entry_date"],
        "Entry ₹": f"₹{float(t['entry'] or 0):,.2f}",
        "SL ₹": f"₹{float(t['stop_loss'] or 0):,.2f}",
        "Target ₹": f"₹{float(t['target'] or 0):,.2f}",
        "Qty": t["quantity"],
        "Confidence": t.get("confidence", "—"),
    } for t in trades]

    return dash_table.DataTable(
        data=rows,
        columns=[{"name": c, "id": c} for c in rows[0].keys()],
        style_table={"overflowX": "auto"},
        style_header={"backgroundColor": "#F9FAFB", "fontWeight": "600",
                      "fontSize": "11.5px", "color": "#6B7280", "border": "none",
                      "textTransform": "uppercase", "letterSpacing": "0.05em"},
        style_cell={"fontSize": "13px", "padding": "9px 12px",
                    "border": "none", "borderBottom": "1px solid #F3F4F6",
                    "fontFamily": "inherit"},
        style_data_conditional=[
            {"if": {"row_index": "odd"}, "backgroundColor": "#FAFAFA"},
        ],
        page_size=10,
    )


def _recent_trades_table() -> html.Div:
    try:
        wb = openpyxl.load_workbook(_EXCEL, data_only=True)
        ws = wb["System"]
        rows = []
        for row in range(19, 50):
            stock = ws.cell(row=row, column=2).value
            if not stock:
                continue
            exit_price = ws.cell(row=row, column=18).value
            entry = float(ws.cell(row=row, column=14).value or 0)
            qty = int(ws.cell(row=row, column=16).value or 0)
            pnl = round((float(exit_price or 0) - entry) * qty, 2) if exit_price else None
            rows.append({
                "Asset": stock,
                "Entry Date": ws.cell(row=row, column=12).value,
                "Entry ₹": f"₹{entry:,.2f}",
                "Exit ₹": f"₹{float(exit_price):,.2f}" if exit_price else "Open",
                "P&L": f"{'▲' if pnl and pnl >= 0 else '▼'} ₹{abs(pnl):,.0f}" if pnl is not None else "—",
                "Status": "Closed" if exit_price else "Open",
            })
    except Exception:
        rows = []

    if not rows:
        return html.Div("No orders. Place a trade via the Paper Trading page.",
                        style={"color": "#9CA3AF", "fontSize": "13px", "padding": "12px 0"})

    return dash_table.DataTable(
        data=rows,
        columns=[{"name": c, "id": c} for c in rows[0].keys()],
        style_table={"overflowX": "auto"},
        style_header={"backgroundColor": "#F9FAFB", "fontWeight": "600",
                      "fontSize": "11.5px", "color": "#6B7280", "border": "none",
                      "textTransform": "uppercase"},
        style_cell={"fontSize": "13px", "padding": "9px 12px",
                    "border": "none", "borderBottom": "1px solid #F3F4F6",
                    "fontFamily": "inherit"},
        style_data_conditional=[
            {"if": {"filter_query": '{Status} = "Open"', "column_id": "Status"},
             "color": "#00A852", "fontWeight": "600"},
            {"if": {"filter_query": '{P&L} contains "▲"', "column_id": "P&L"},
             "color": "#00A852"},
            {"if": {"filter_query": '{P&L} contains "▼"', "column_id": "P&L"},
             "color": "#E03131"},
        ],
        page_size=10,
    )


# ── Callbacks ─────────────────────────────────────────────────────────────

@callback(
    Output("portfolio-chart", "figure"),
    Output("tf-store", "data"),
    Input("tf-1d", "n_clicks"), Input("tf-1m", "n_clicks"),
    Input("tf-1y", "n_clicks"), Input("tf-all", "n_clicks"),
    Input("tf-refresh", "n_clicks"),
    prevent_initial_call=True,
)
def update_chart(n1d, n1m, n1y, nall, nref):
    from dash import ctx
    tid = ctx.triggered_id or "tf-all"
    period = {"tf-1d": "1D", "tf-1m": "1M", "tf-1y": "1Y"}.get(tid, "All")
    return _portfolio_chart(period), period


@callback(
    Output("positions-table-div", "children"),
    Output("recent-trades-div", "children"),
    Input("home-refresh", "n_intervals"),
    Input("exec-exits-btn", "n_clicks"),
    prevent_initial_call=False,
)
def refresh_tables(_, n_clicks):
    return _positions_table(), _recent_trades_table()
